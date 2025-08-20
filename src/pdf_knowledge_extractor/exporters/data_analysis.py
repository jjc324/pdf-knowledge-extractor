"""
Data & Analysis Export Formats
Support for advanced CSV, Excel with analytics, JSON-LD, and RDF semantic formats
"""

import logging
import json
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import time
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
import re

from .base import BaseExporter, ExportConfig, ExportResult

logger = logging.getLogger(__name__)


class AdvancedCSVExporter(BaseExporter):
    """Export to comprehensive CSV format with multiple tables."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['advanced-csv', 'csv-multi']
    
    @property
    def file_extension(self) -> str:
        return '.zip'  # Multiple CSV files in a zip
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to multiple CSV files with comprehensive data."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            # Create output directory
            output_dir = self.config.output_path
            if output_dir.suffix == '.zip':
                output_dir = output_dir.parent / output_dir.stem
            
            output_dir.mkdir(parents=True, exist_ok=True)
            
            semantic_data = analysis_data.get('semantic_analysis', {})
            exported_docs = list(documents.keys())
            total_concepts = len(semantic_data.get('concepts', []))
            
            # Create documents CSV
            self._create_documents_csv(output_dir, analysis_data, documents)
            
            # Create concepts CSV
            if semantic_data.get('concepts'):
                self._create_concepts_csv(output_dir, semantic_data['concepts'])
            
            # Create relationships CSV
            if semantic_data.get('similarities'):
                self._create_relationships_csv(output_dir, semantic_data['similarities'])
            
            # Create clusters CSV
            if semantic_data.get('clusters'):
                self._create_clusters_csv(output_dir, semantic_data['clusters'])
            
            # Create concept-document mapping CSV
            if semantic_data.get('concepts'):
                self._create_concept_document_csv(output_dir, semantic_data['concepts'])
            
            # Create summary statistics CSV
            self._create_statistics_csv(output_dir, analysis_data, documents)
            
            # Create README
            self._create_readme(output_dir, analysis_data, documents)
            
            # Zip the directory if requested
            final_output_path = output_dir
            if self.config.output_path.suffix == '.zip':
                import shutil
                shutil.make_archive(str(self.config.output_path.with_suffix('')), 'zip', output_dir)
                final_output_path = self.config.output_path
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=final_output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                stats={'csv_files': len(list(output_dir.glob('*.csv')))}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Advanced CSV export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_documents_csv(self, output_dir: Path, analysis_data: Dict, documents: Dict):
        """Create documents overview CSV."""
        csv_path = output_dir / "documents.csv"
        
        headers = [
            'document_id', 'filename', 'word_count', 'character_count', 'sentiment_score', 
            'sentiment_label', 'topic_count', 'concept_count', 'cluster_id', 'cluster_label',
            'analysis_date'
        ]
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for doc_id, doc_text in documents.items():
                doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
                
                # Get concept count for this document
                concept_count = 0
                if semantic_data and semantic_data.get('concepts'):
                    concept_count = len([c for c in semantic_data['concepts'] 
                                       if doc_id in c.get('document_ids', [])])
                
                # Get cluster information
                cluster_id = ""
                cluster_label = ""
                if semantic_data and semantic_data.get('clusters'):
                    for cluster in semantic_data['clusters']:
                        if doc_id in cluster.get('document_ids', []):
                            cluster_id = cluster.get('cluster_id', '')
                            cluster_label = cluster.get('cluster_label', '')
                            break
                
                # Get sentiment
                sentiment = doc_analysis.get('sentiment', {})
                
                writer.writerow([
                    doc_id,
                    doc_id,  # filename same as doc_id
                    doc_analysis.get('word_count', 0),
                    doc_analysis.get('character_count', 0),
                    sentiment.get('score', 0),
                    sentiment.get('sentiment', 'neutral'),
                    len(doc_analysis.get('topics', [])),
                    concept_count,
                    cluster_id,
                    cluster_label,
                    datetime.now().isoformat()
                ])
    
    def _create_concepts_csv(self, output_dir: Path, concepts: List[Dict]):
        """Create concepts CSV."""
        csv_path = output_dir / "concepts.csv"
        
        headers = [
            'concept_id', 'concept_text', 'concept_type', 'importance_score', 
            'frequency', 'document_count', 'context_sentence', 'first_document'
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for i, concept in enumerate(concepts):
                context = concept.get('context_sentences', [''])[0][:200] if concept.get('context_sentences') else ''
                first_doc = concept.get('document_ids', [''])[0]
                
                writer.writerow([
                    f"concept_{i+1}",
                    concept.get('text', ''),
                    concept.get('concept_type', 'unknown'),
                    concept.get('importance_score', 0),
                    concept.get('frequency', 0),
                    len(concept.get('document_ids', [])),
                    context,
                    first_doc
                ])
    
    def _create_relationships_csv(self, output_dir: Path, similarities: List[Dict]):
        """Create document relationships CSV."""
        csv_path = output_dir / "relationships.csv"
        
        headers = [
            'relationship_id', 'document_1', 'document_2', 'similarity_score', 
            'similarity_type', 'shared_concepts', 'relationship_strength'
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for i, sim in enumerate(similarities):
                similarity_score = sim.get('similarity_score', 0)
                strength = 'weak' if similarity_score < 0.3 else 'moderate' if similarity_score < 0.7 else 'strong'
                shared_concepts = '; '.join(sim.get('shared_concepts', []))
                
                writer.writerow([
                    f"rel_{i+1}",
                    sim.get('doc1_id', ''),
                    sim.get('doc2_id', ''),
                    similarity_score,
                    sim.get('similarity_type', 'cosine'),
                    shared_concepts,
                    strength
                ])
    
    def _create_clusters_csv(self, output_dir: Path, clusters: List[Dict]):
        """Create clusters CSV."""
        csv_path = output_dir / "clusters.csv"
        
        headers = [
            'cluster_id', 'cluster_label', 'document_count', 'coherence_score',
            'main_topics', 'document_list'
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for cluster in clusters:
                doc_ids = cluster.get('document_ids', [])
                topics = '; '.join(cluster.get('main_topics', []))
                doc_list = '; '.join(doc_ids)
                
                writer.writerow([
                    cluster.get('cluster_id', ''),
                    cluster.get('cluster_label', ''),
                    len(doc_ids),
                    cluster.get('coherence_score', 0),
                    topics,
                    doc_list
                ])
    
    def _create_concept_document_csv(self, output_dir: Path, concepts: List[Dict]):
        """Create concept-document mapping CSV."""
        csv_path = output_dir / "concept_document_mapping.csv"
        
        headers = ['concept_text', 'document_id', 'importance_score', 'frequency_in_doc']
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for concept in concepts:
                concept_text = concept.get('text', '')
                importance = concept.get('importance_score', 0)
                
                for doc_id in concept.get('document_ids', []):
                    writer.writerow([
                        concept_text,
                        doc_id,
                        importance,
                        concept.get('frequency', 0)  # This is global frequency, could be per-doc
                    ])
    
    def _create_statistics_csv(self, output_dir: Path, analysis_data: Dict, documents: Dict):
        """Create summary statistics CSV."""
        csv_path = output_dir / "statistics.csv"
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        stats = [
            ('total_documents', len(documents)),
            ('total_words', sum(analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                              for doc_id in documents.keys())),
            ('total_concepts', len(semantic_data.get('concepts', []))),
            ('total_relationships', len(semantic_data.get('similarities', []))),
            ('total_clusters', len(semantic_data.get('clusters', []))),
            ('avg_words_per_doc', sum(analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                                    for doc_id in documents.keys()) / len(documents) if documents else 0),
            ('avg_concepts_per_doc', len(semantic_data.get('concepts', [])) / len(documents) if documents else 0),
            ('analysis_date', datetime.now().isoformat())
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['statistic', 'value'])
            writer.writerows(stats)
    
    def _create_readme(self, output_dir: Path, analysis_data: Dict, documents: Dict):
        """Create README file explaining the CSV structure."""
        readme_path = output_dir / "README.md"
        
        content = f"""# PDF Knowledge Analysis CSV Export

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Total documents: {len(documents)}
Total concepts: {len(analysis_data.get('semantic_analysis', {}).get('concepts', []))}

## File Descriptions

### documents.csv
Contains metadata and analysis results for each document:
- document_id: Unique identifier for the document
- word_count: Total words in the document
- concept_count: Number of concepts extracted from this document
- cluster_id/cluster_label: Cluster assignment (if applicable)

### concepts.csv
Contains all extracted concepts:
- concept_text: The actual concept text
- concept_type: Type of concept (entity, keyword, topic, etc.)
- importance_score: Importance score (0-1)
- document_count: Number of documents containing this concept

### relationships.csv
Contains document similarity relationships:
- document_1/document_2: The two related documents
- similarity_score: Cosine similarity score (0-1)
- shared_concepts: Concepts shared between documents

### clusters.csv
Contains document cluster information:
- cluster_label: Human-readable cluster name
- document_count: Number of documents in cluster
- main_topics: Key topics for this cluster

### concept_document_mapping.csv
Contains mapping between concepts and documents:
- Links each concept to the documents it appears in
- Useful for network analysis and concept co-occurrence

### statistics.csv
Contains summary statistics about the entire analysis.

## Usage Tips

1. Load documents.csv first to understand the dataset structure
2. Use concept_document_mapping.csv for network analysis
3. Join tables using document_id as the key
4. Filter concepts by importance_score for focused analysis
"""
        
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write(content)


class AdvancedExcelExporter(BaseExporter):
    """Export to Excel with advanced charts, pivot tables, and analytics."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['advanced-excel', 'excel-analytics']
    
    @property
    def file_extension(self) -> str:
        return '.xlsx'
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to Excel with advanced analytics."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            semantic_data = analysis_data.get('semantic_analysis', {})
            exported_docs = list(documents.keys())
            total_concepts = len(semantic_data.get('concepts', []))
            
            # Create Dashboard sheet
            self._create_dashboard_sheet(workbook, analysis_data, documents)
            
            # Create Documents sheet with advanced formatting
            self._create_documents_sheet_advanced(workbook, analysis_data, documents)
            
            # Create Concepts sheet
            if semantic_data.get('concepts'):
                self._create_concepts_sheet_advanced(workbook, semantic_data['concepts'])
            
            # Create Analytics sheet
            self._create_analytics_sheet(workbook, analysis_data, documents)
            
            # Create Relationships sheet with network analysis
            if semantic_data.get('similarities'):
                self._create_relationships_sheet_advanced(workbook, semantic_data['similarities'])
            
            # Create Charts sheet
            self._create_charts_sheet(workbook, analysis_data, documents)
            
            # Make Dashboard the first sheet
            dashboard = workbook['Dashboard']
            workbook.move_sheet(dashboard, offset=-len(workbook.worksheets))
            
            # Save workbook
            workbook.save(output_path)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                stats={'excel_sheets': len(workbook.worksheets)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Advanced Excel export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_dashboard_sheet(self, workbook, analysis_data: Dict, documents: Dict):
        """Create executive dashboard sheet."""
        dashboard = workbook.create_sheet("Dashboard")
        
        # Title
        dashboard['A1'] = 'PDF Knowledge Analysis Dashboard'
        dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        dashboard['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        dashboard.merge_cells('A1:F1')
        
        # Key metrics
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        metrics = [
            ('Total Documents', len(documents)),
            ('Total Concepts', len(semantic_data.get('concepts', []))),
            ('Document Relationships', len(semantic_data.get('similarities', []))),
            ('Document Clusters', len(semantic_data.get('clusters', []))),
            ('Total Words', sum(analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                              for doc_id in documents.keys())),
            ('Analysis Date', datetime.now().strftime('%Y-%m-%d'))
        ]
        
        row = 3
        for metric, value in metrics:
            dashboard[f'A{row}'] = metric
            dashboard[f'B{row}'] = value
            dashboard[f'A{row}'].font = Font(bold=True)
            dashboard[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row += 1
        
        # Top concepts
        dashboard['D3'] = 'Top 10 Concepts'
        dashboard['D3'].font = Font(bold=True, size=12)
        
        if semantic_data.get('concepts'):
            sorted_concepts = sorted(semantic_data['concepts'], 
                                   key=lambda x: x.get('importance_score', 0), reverse=True)
            
            dashboard['D4'] = 'Concept'
            dashboard['E4'] = 'Importance'
            dashboard['F4'] = 'Frequency'
            
            for i, concept in enumerate(sorted_concepts[:10], 5):
                dashboard[f'D{i}'] = concept.get('text', '')
                dashboard[f'E{i}'] = concept.get('importance_score', 0)
                dashboard[f'F{i}'] = concept.get('frequency', 0)
        
        # Auto-adjust column widths
        for column in dashboard.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 40)
            dashboard.column_dimensions[column_letter].width = adjusted_width
    
    def _create_documents_sheet_advanced(self, workbook, analysis_data: Dict, documents: Dict):
        """Create advanced documents sheet with conditional formatting."""
        doc_sheet = workbook.create_sheet("Documents")
        
        headers = [
            'Document ID', 'Word Count', 'Character Count', 'Concepts', 'Sentiment Score',
            'Sentiment', 'Topics', 'Cluster', 'Related Docs', 'Analysis Date'
        ]
        
        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = doc_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        # Add data rows
        for row, (doc_id, doc_text) in enumerate(documents.items(), 2):
            doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
            
            # Get concept count
            concept_count = 0
            if semantic_data and semantic_data.get('concepts'):
                concept_count = len([c for c in semantic_data['concepts'] 
                                   if doc_id in c.get('document_ids', [])])
            
            # Get cluster
            cluster_label = ""
            if semantic_data and semantic_data.get('clusters'):
                for cluster in semantic_data['clusters']:
                    if doc_id in cluster.get('document_ids', []):
                        cluster_label = cluster.get('cluster_label', 'Unnamed')
                        break
            
            # Get related docs count
            related_count = 0
            if semantic_data and semantic_data.get('similarities'):
                for sim in semantic_data['similarities']:
                    if (sim.get('doc1_id') == doc_id or sim.get('doc2_id') == doc_id) and \
                       sim.get('similarity_score', 0) >= 0.5:
                        related_count += 1
            
            # Get sentiment
            sentiment = doc_analysis.get('sentiment', {})
            
            # Get topics
            topics = doc_analysis.get('topics', [])
            top_topics = ', '.join([t.get('topic', '') for t in topics[:3]])
            
            values = [
                doc_id,
                doc_analysis.get('word_count', 0),
                doc_analysis.get('character_count', 0),
                concept_count,
                sentiment.get('score', 0),
                sentiment.get('sentiment', 'neutral'),
                top_topics,
                cluster_label,
                related_count,
                datetime.now().strftime('%Y-%m-%d')
            ]
            
            for col, value in enumerate(values, 1):
                doc_sheet.cell(row=row, column=col, value=value)
        
        # Add conditional formatting for sentiment scores
        from openpyxl.formatting.rule import ColorScaleRule
        sentiment_col = 'E2:E' + str(len(documents) + 1)
        color_rule = ColorScaleRule(start_type='min', start_color='FF6B6B',
                                  mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                                  end_type='max', end_color='4CAF50')
        doc_sheet.conditional_formatting.add(sentiment_col, color_rule)
        
        # Auto-adjust column widths
        for column in doc_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 40)
            doc_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_concepts_sheet_advanced(self, workbook, concepts: List[Dict]):
        """Create advanced concepts sheet."""
        concepts_sheet = workbook.create_sheet("Concepts")
        
        headers = ['Concept', 'Type', 'Importance', 'Frequency', 'Documents', 'Context Preview']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = concepts_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sort concepts by importance
        sorted_concepts = sorted(concepts, key=lambda x: x.get('importance_score', 0), reverse=True)
        
        # Add data
        for row, concept in enumerate(sorted_concepts, 2):
            context = concept.get('context_sentences', [''])[0][:50] if concept.get('context_sentences') else ''
            if len(concept.get('context_sentences', [''])[0]) > 50:
                context += "..."
            
            values = [
                concept.get('text', ''),
                concept.get('concept_type', 'unknown'),
                concept.get('importance_score', 0),
                concept.get('frequency', 0),
                len(concept.get('document_ids', [])),
                context
            ]
            
            for col, value in enumerate(values, 1):
                concepts_sheet.cell(row=row, column=col, value=value)
        
        # Add conditional formatting for importance
        from openpyxl.formatting.rule import ColorScaleRule
        importance_col = f'C2:C{len(sorted_concepts) + 1}'
        color_rule = ColorScaleRule(start_type='min', start_color='FFCDD2',
                                  end_type='max', end_color='C8E6C9')
        concepts_sheet.conditional_formatting.add(importance_col, color_rule)
        
        # Auto-adjust columns
        for column in concepts_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            concepts_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_analytics_sheet(self, workbook, analysis_data: Dict, documents: Dict):
        """Create analytics sheet with pivot-like summaries."""
        analytics = workbook.create_sheet("Analytics")
        
        # Title
        analytics['A1'] = 'Document Collection Analytics'
        analytics['A1'].font = Font(size=16, bold=True)
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        # Word count distribution
        analytics['A3'] = 'Word Count Distribution'
        analytics['A3'].font = Font(bold=True)
        
        word_counts = [analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                      for doc_id in documents.keys()]
        
        if word_counts:
            analytics['A4'] = 'Minimum'
            analytics['B4'] = min(word_counts)
            analytics['A5'] = 'Maximum'
            analytics['B5'] = max(word_counts)
            analytics['A6'] = 'Average'
            analytics['B6'] = sum(word_counts) / len(word_counts)
            analytics['A7'] = 'Median'
            analytics['B7'] = sorted(word_counts)[len(word_counts)//2]
        
        # Concept type distribution
        if semantic_data.get('concepts'):
            analytics['D3'] = 'Concept Type Distribution'
            analytics['D3'].font = Font(bold=True)
            
            concept_types = {}
            for concept in semantic_data['concepts']:
                concept_type = concept.get('concept_type', 'unknown')
                concept_types[concept_type] = concept_types.get(concept_type, 0) + 1
            
            row = 4
            for concept_type, count in concept_types.items():
                analytics[f'D{row}'] = concept_type
                analytics[f'E{row}'] = count
                row += 1
        
        # Sentiment distribution
        analytics['A10'] = 'Sentiment Distribution'
        analytics['A10'].font = Font(bold=True)
        
        sentiments = {}
        for doc_id in documents.keys():
            doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
            sentiment = doc_analysis.get('sentiment', {}).get('sentiment', 'neutral')
            sentiments[sentiment] = sentiments.get(sentiment, 0) + 1
        
        row = 11
        for sentiment, count in sentiments.items():
            analytics[f'A{row}'] = sentiment
            analytics[f'B{row}'] = count
            row += 1
    
    def _create_relationships_sheet_advanced(self, workbook, similarities: List[Dict]):
        """Create advanced relationships sheet with network metrics."""
        rel_sheet = workbook.create_sheet("Relationships")
        
        headers = ['Document 1', 'Document 2', 'Similarity', 'Strength', 'Shared Concepts']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = rel_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Sort by similarity
        sorted_sims = sorted(similarities, key=lambda x: x.get('similarity_score', 0), reverse=True)
        
        # Add data
        for row, sim in enumerate(sorted_sims, 2):
            similarity = sim.get('similarity_score', 0)
            strength = 'Weak' if similarity < 0.3 else 'Moderate' if similarity < 0.7 else 'Strong'
            shared_concepts = ', '.join(sim.get('shared_concepts', [])[:5])
            
            values = [
                sim.get('doc1_id', ''),
                sim.get('doc2_id', ''),
                similarity,
                strength,
                shared_concepts
            ]
            
            for col, value in enumerate(values, 1):
                rel_sheet.cell(row=row, column=col, value=value)
        
        # Add conditional formatting for similarity scores
        from openpyxl.formatting.rule import ColorScaleRule
        sim_col = f'C2:C{len(sorted_sims) + 1}'
        color_rule = ColorScaleRule(start_type='min', start_color='FFCDD2',
                                  end_type='max', end_color='C8E6C9')
        rel_sheet.conditional_formatting.add(sim_col, color_rule)
        
        # Auto-adjust columns
        for column in rel_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 40)
            rel_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_charts_sheet(self, workbook, analysis_data: Dict, documents: Dict):
        """Create charts and visualizations sheet."""
        charts_sheet = workbook.create_sheet("Charts")
        
        # Title
        charts_sheet['A1'] = 'Data Visualizations'
        charts_sheet['A1'].font = Font(size=16, bold=True)
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        # Create data for word count chart
        doc_names = list(documents.keys())[:10]  # Top 10 documents
        word_counts = [analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                      for doc_id in doc_names]
        
        # Add word count data
        charts_sheet['A3'] = 'Document'
        charts_sheet['B3'] = 'Word Count'
        
        for i, (doc_name, word_count) in enumerate(zip(doc_names, word_counts), 4):
            charts_sheet[f'A{i}'] = doc_name[:20]  # Truncate long names
            charts_sheet[f'B{i}'] = word_count
        
        # Create bar chart for word counts
        chart = BarChart()
        chart.title = "Word Count by Document"
        chart.x_axis.title = "Documents"
        chart.y_axis.title = "Word Count"
        
        data = Reference(charts_sheet, min_col=2, min_row=3, max_row=3 + len(doc_names))
        categories = Reference(charts_sheet, min_col=1, min_row=4, max_row=3 + len(doc_names))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        charts_sheet.add_chart(chart, "D3")
        
        # Create concept type pie chart if concepts exist
        if semantic_data.get('concepts'):
            concept_types = {}
            for concept in semantic_data['concepts']:
                concept_type = concept.get('concept_type', 'unknown')
                concept_types[concept_type] = concept_types.get(concept_type, 0) + 1
            
            # Add concept type data
            start_row = 15
            charts_sheet[f'A{start_row}'] = 'Concept Type'
            charts_sheet[f'B{start_row}'] = 'Count'
            
            for i, (concept_type, count) in enumerate(concept_types.items(), start_row + 1):
                charts_sheet[f'A{i}'] = concept_type
                charts_sheet[f'B{i}'] = count
            
            # Create pie chart
            pie = PieChart()
            pie.title = "Concept Type Distribution"
            
            data = Reference(charts_sheet, min_col=2, min_row=start_row, 
                           max_row=start_row + len(concept_types))
            categories = Reference(charts_sheet, min_col=1, min_row=start_row + 1, 
                                 max_row=start_row + len(concept_types))
            
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(categories)
            charts_sheet.add_chart(pie, "D15")


class JSONLDExporter(BaseExporter):
    """Export to JSON-LD structured data format."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['json-ld', 'jsonld']
    
    @property
    def file_extension(self) -> str:
        return '.jsonld'
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to JSON-LD structured data format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create JSON-LD structure
            jsonld_data = self._create_jsonld_structure(analysis_data, documents)
            
            # Write JSON-LD file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(jsonld_data, f, indent=2, ensure_ascii=False)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=list(documents.keys()),
                execution_time=execution_time,
                exported_concepts=len(analysis_data.get('semantic_analysis', {}).get('concepts', [])),
                stats={'jsonld_objects': len(jsonld_data.get('@graph', []))}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"JSON-LD export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_jsonld_structure(self, analysis_data: Dict, documents: Dict) -> Dict[str, Any]:
        """Create JSON-LD structured data."""
        
        base_uri = "https://pdfextractor.ai/kb/"
        
        jsonld = {
            "@context": {
                "@vocab": "https://schema.org/",
                "pdf": "https://pdfextractor.ai/ontology/",
                "dct": "http://purl.org/dc/terms/",
                "skos": "http://www.w3.org/2004/02/skos/core#",
                "foaf": "http://xmlns.com/foaf/0.1/",
                "rdfs": "http://www.w3.org/2000/01/rdf-schema#"
            },
            "@id": base_uri + "knowledge-base",
            "@type": "Dataset",
            "name": "PDF Knowledge Base",
            "description": "Knowledge extracted from PDF documents using semantic analysis",
            "dateCreated": datetime.now().isoformat(),
            "creator": {
                "@type": "SoftwareApplication",
                "name": "PDF Knowledge Extractor",
                "version": "v2.2"
            },
            "@graph": []
        }
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        # Add documents
        for doc_id, doc_text in documents.items():
            doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
            doc_obj = self._create_document_jsonld(doc_id, doc_text, doc_analysis, base_uri)
            jsonld["@graph"].append(doc_obj)
        
        # Add concepts
        if semantic_data.get('concepts'):
            for concept in semantic_data['concepts']:
                concept_obj = self._create_concept_jsonld(concept, base_uri)
                jsonld["@graph"].append(concept_obj)
        
        # Add relationships
        if semantic_data.get('similarities'):
            for i, sim in enumerate(semantic_data['similarities']):
                rel_obj = self._create_relationship_jsonld(sim, i, base_uri)
                jsonld["@graph"].append(rel_obj)
        
        # Add clusters
        if semantic_data.get('clusters'):
            for cluster in semantic_data['clusters']:
                cluster_obj = self._create_cluster_jsonld(cluster, base_uri)
                jsonld["@graph"].append(cluster_obj)
        
        return jsonld
    
    def _create_document_jsonld(self, doc_id: str, doc_text: str, doc_analysis: Dict, base_uri: str) -> Dict[str, Any]:
        """Create JSON-LD object for a document."""
        doc_uri = base_uri + "document/" + self._uri_encode(doc_id)
        
        doc_obj = {
            "@id": doc_uri,
            "@type": ["DigitalDocument", "pdf:AnalyzedDocument"],
            "name": doc_id,
            "dct:title": doc_id,
            "wordCount": doc_analysis.get('word_count', 0),
            "characterCount": doc_analysis.get('character_count', 0),
            "dateAnalyzed": datetime.now().isoformat(),
            "pdf:sentimentScore": doc_analysis.get('sentiment', {}).get('score', 0),
            "pdf:sentimentLabel": doc_analysis.get('sentiment', {}).get('sentiment', 'neutral'),
        }
        
        # Add topics
        topics = doc_analysis.get('topics', [])
        if topics:
            doc_obj["about"] = []
            for topic in topics[:5]:  # Top 5 topics
                topic_obj = {
                    "@type": "DefinedTerm",
                    "name": topic.get('topic', ''),
                    "pdf:frequency": topic.get('frequency', 0) if 'frequency' in topic else None
                }
                # Remove None values
                topic_obj = {k: v for k, v in topic_obj.items() if v is not None}
                doc_obj["about"].append(topic_obj)
        
        return doc_obj
    
    def _create_concept_jsonld(self, concept: Dict, base_uri: str) -> Dict[str, Any]:
        """Create JSON-LD object for a concept."""
        concept_uri = base_uri + "concept/" + self._uri_encode(concept.get('text', ''))
        
        concept_obj = {
            "@id": concept_uri,
            "@type": ["DefinedTerm", "skos:Concept"],
            "name": concept.get('text', ''),
            "skos:prefLabel": concept.get('text', ''),
            "pdf:conceptType": concept.get('concept_type', 'unknown'),
            "pdf:importanceScore": concept.get('importance_score', 0),
            "pdf:frequency": concept.get('frequency', 0),
            "pdf:documentCount": len(concept.get('document_ids', []))
        }
        
        # Add context if available
        if concept.get('context_sentences'):
            concept_obj["description"] = concept['context_sentences'][0][:200]
        
        # Add related documents
        if concept.get('document_ids'):
            concept_obj["pdf:appearsIn"] = []
            for doc_id in concept['document_ids']:
                doc_uri = base_uri + "document/" + self._uri_encode(doc_id)
                concept_obj["pdf:appearsIn"].append({"@id": doc_uri})
        
        return concept_obj
    
    def _create_relationship_jsonld(self, similarity: Dict, index: int, base_uri: str) -> Dict[str, Any]:
        """Create JSON-LD object for a document relationship."""
        rel_uri = base_uri + "relationship/" + str(index)
        
        doc1_uri = base_uri + "document/" + self._uri_encode(similarity.get('doc1_id', ''))
        doc2_uri = base_uri + "document/" + self._uri_encode(similarity.get('doc2_id', ''))
        
        rel_obj = {
            "@id": rel_uri,
            "@type": "pdf:DocumentSimilarity",
            "pdf:document1": {"@id": doc1_uri},
            "pdf:document2": {"@id": doc2_uri},
            "pdf:similarityScore": similarity.get('similarity_score', 0),
            "pdf:similarityType": similarity.get('similarity_type', 'cosine')
        }
        
        # Add shared concepts
        if similarity.get('shared_concepts'):
            rel_obj["pdf:sharedConcepts"] = []
            for concept_text in similarity['shared_concepts'][:10]:  # Max 10
                concept_uri = base_uri + "concept/" + self._uri_encode(concept_text)
                rel_obj["pdf:sharedConcepts"].append({"@id": concept_uri})
        
        return rel_obj
    
    def _create_cluster_jsonld(self, cluster: Dict, base_uri: str) -> Dict[str, Any]:
        """Create JSON-LD object for a document cluster."""
        cluster_uri = base_uri + "cluster/" + self._uri_encode(cluster.get('cluster_id', ''))
        
        cluster_obj = {
            "@id": cluster_uri,
            "@type": "pdf:DocumentCluster",
            "name": cluster.get('cluster_label', 'Unnamed Cluster'),
            "pdf:coherenceScore": cluster.get('coherence_score', 0),
            "pdf:documentCount": len(cluster.get('document_ids', []))
        }
        
        # Add member documents
        if cluster.get('document_ids'):
            cluster_obj["pdf:contains"] = []
            for doc_id in cluster['document_ids']:
                doc_uri = base_uri + "document/" + self._uri_encode(doc_id)
                cluster_obj["pdf:contains"].append({"@id": doc_uri})
        
        # Add main topics
        if cluster.get('main_topics'):
            cluster_obj["about"] = []
            for topic in cluster['main_topics'][:5]:  # Top 5 topics
                cluster_obj["about"].append({
                    "@type": "DefinedTerm",
                    "name": topic
                })
        
        return cluster_obj
    
    def _uri_encode(self, text: str) -> str:
        """Encode text for use in URIs."""
        # Simple URI encoding - replace spaces and special characters
        encoded = re.sub(r'[^\w\-.]', '_', text)
        encoded = re.sub(r'_{2,}', '_', encoded)  # Multiple underscores to single
        return encoded.strip('_')[:100]  # Limit length


class RDFExporter(BaseExporter):
    """Export to RDF/XML semantic web format."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['rdf', 'rdf-xml']
    
    @property
    def file_extension(self) -> str:
        return '.rdf'
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to RDF/XML format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create RDF XML structure
            rdf_root = self._create_rdf_xml(analysis_data, documents)
            
            # Write RDF file
            rough_string = ET.tostring(rdf_root, encoding='unicode')
            reparsed = minidom.parseString(rough_string)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(reparsed.toprettyxml(indent="  "))
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=list(documents.keys()),
                execution_time=execution_time,
                exported_concepts=len(analysis_data.get('semantic_analysis', {}).get('concepts', [])),
                stats={'rdf_triples': len(list(rdf_root))}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"RDF export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_rdf_xml(self, analysis_data: Dict, documents: Dict) -> ET.Element:
        """Create RDF/XML structure."""
        
        # Define namespaces
        namespaces = {
            'rdf': 'http://www.w3.org/1999/02/22-rdf-syntax-ns#',
            'rdfs': 'http://www.w3.org/2000/01/rdf-schema#',
            'dc': 'http://purl.org/dc/elements/1.1/',
            'dct': 'http://purl.org/dc/terms/',
            'foaf': 'http://xmlns.com/foaf/0.1/',
            'skos': 'http://www.w3.org/2004/02/skos/core#',
            'pdf': 'https://pdfextractor.ai/ontology/'
        }
        
        # Create root element
        root = ET.Element('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}RDF')
        for prefix, uri in namespaces.items():
            root.set(f'xmlns:{prefix}', uri)
        
        base_uri = "https://pdfextractor.ai/kb/"
        semantic_data = analysis_data.get('semantic_analysis', {})
        
        # Add documents
        for doc_id, doc_text in documents.items():
            doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
            doc_elem = self._create_document_rdf(root, doc_id, doc_text, doc_analysis, base_uri, namespaces)
            root.append(doc_elem)
        
        # Add concepts
        if semantic_data.get('concepts'):
            for concept in semantic_data['concepts']:
                concept_elem = self._create_concept_rdf(root, concept, base_uri, namespaces)
                root.append(concept_elem)
        
        # Add relationships
        if semantic_data.get('similarities'):
            for i, sim in enumerate(semantic_data['similarities']):
                rel_elem = self._create_relationship_rdf(root, sim, i, base_uri, namespaces)
                root.append(rel_elem)
        
        return root
    
    def _create_document_rdf(self, root, doc_id: str, doc_text: str, doc_analysis: Dict, 
                           base_uri: str, namespaces: Dict) -> ET.Element:
        """Create RDF element for a document."""
        doc_uri = base_uri + "document/" + self._uri_encode(doc_id)
        
        doc_elem = ET.Element('{https://pdfextractor.ai/ontology/}AnalyzedDocument')
        doc_elem.set('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}about', doc_uri)
        
        # Add properties
        title_elem = ET.SubElement(doc_elem, '{http://purl.org/dc/elements/1.1/}title')
        title_elem.text = doc_id
        
        word_count_elem = ET.SubElement(doc_elem, '{https://pdfextractor.ai/ontology/}wordCount')
        word_count_elem.text = str(doc_analysis.get('word_count', 0))
        
        char_count_elem = ET.SubElement(doc_elem, '{https://pdfextractor.ai/ontology/}characterCount')
        char_count_elem.text = str(doc_analysis.get('character_count', 0))
        
        # Add sentiment
        sentiment = doc_analysis.get('sentiment', {})
        if sentiment:
            sentiment_score_elem = ET.SubElement(doc_elem, '{https://pdfextractor.ai/ontology/}sentimentScore')
            sentiment_score_elem.text = str(sentiment.get('score', 0))
            
            sentiment_label_elem = ET.SubElement(doc_elem, '{https://pdfextractor.ai/ontology/}sentimentLabel')
            sentiment_label_elem.text = sentiment.get('sentiment', 'neutral')
        
        # Add creation date
        created_elem = ET.SubElement(doc_elem, '{http://purl.org/dc/terms/}created')
        created_elem.text = datetime.now().isoformat()
        
        return doc_elem
    
    def _create_concept_rdf(self, root, concept: Dict, base_uri: str, namespaces: Dict) -> ET.Element:
        """Create RDF element for a concept."""
        concept_uri = base_uri + "concept/" + self._uri_encode(concept.get('text', ''))
        
        concept_elem = ET.Element('{http://www.w3.org/2004/02/skos/core#}Concept')
        concept_elem.set('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}about', concept_uri)
        
        # Add properties
        pref_label_elem = ET.SubElement(concept_elem, '{http://www.w3.org/2004/02/skos/core#}prefLabel')
        pref_label_elem.text = concept.get('text', '')
        
        concept_type_elem = ET.SubElement(concept_elem, '{https://pdfextractor.ai/ontology/}conceptType')
        concept_type_elem.text = concept.get('concept_type', 'unknown')
        
        importance_elem = ET.SubElement(concept_elem, '{https://pdfextractor.ai/ontology/}importanceScore')
        importance_elem.text = str(concept.get('importance_score', 0))
        
        frequency_elem = ET.SubElement(concept_elem, '{https://pdfextractor.ai/ontology/}frequency')
        frequency_elem.text = str(concept.get('frequency', 0))
        
        # Add context if available
        if concept.get('context_sentences'):
            definition_elem = ET.SubElement(concept_elem, '{http://www.w3.org/2004/02/skos/core#}definition')
            definition_elem.text = concept['context_sentences'][0][:200]
        
        return concept_elem
    
    def _create_relationship_rdf(self, root, similarity: Dict, index: int, 
                               base_uri: str, namespaces: Dict) -> ET.Element:
        """Create RDF element for a relationship."""
        rel_uri = base_uri + "relationship/" + str(index)
        
        rel_elem = ET.Element('{https://pdfextractor.ai/ontology/}DocumentSimilarity')
        rel_elem.set('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}about', rel_uri)
        
        # Add document references
        doc1_uri = base_uri + "document/" + self._uri_encode(similarity.get('doc1_id', ''))
        doc2_uri = base_uri + "document/" + self._uri_encode(similarity.get('doc2_id', ''))
        
        doc1_elem = ET.SubElement(rel_elem, '{https://pdfextractor.ai/ontology/}document1')
        doc1_elem.set('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}resource', doc1_uri)
        
        doc2_elem = ET.SubElement(rel_elem, '{https://pdfextractor.ai/ontology/}document2')
        doc2_elem.set('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}resource', doc2_uri)
        
        # Add similarity score
        score_elem = ET.SubElement(rel_elem, '{https://pdfextractor.ai/ontology/}similarityScore')
        score_elem.text = str(similarity.get('similarity_score', 0))
        
        return rel_elem
    
    def _uri_encode(self, text: str) -> str:
        """Encode text for use in URIs."""
        encoded = re.sub(r'[^\w\-.]', '_', text)
        encoded = re.sub(r'_{2,}', '_', encoded)
        return encoded.strip('_')[:100]