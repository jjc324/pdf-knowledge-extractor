"""
Academic & Research Export Formats
Support for Zotero, LaTeX, Google Docs, and structured data formats
"""

import logging
import json
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import time
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re

from .base import BaseExporter, TemplateExporter, ExportConfig, ExportResult

logger = logging.getLogger(__name__)


class ZoteroExporter(TemplateExporter):
    """Export to Zotero library format with metadata, notes, and collections."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['zotero']
    
    @property
    def file_extension(self) -> str:
        return '.rdf'
    
    @property
    def default_templates(self) -> Dict[str, str]:
        return {
            'rdf_header': """<?xml version="1.0" encoding="UTF-8"?>
<rdf:RDF
    xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
    xmlns:z="http://www.zotero.org/namespaces/export#"
    xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:dcterms="http://purl.org/dc/terms/"
    xmlns:bib="http://purl.org/net/biblio#"
    xmlns:foaf="http://xmlns.com/foaf/0.1/">""",
            'document_item': """
    <bib:Document rdf:about="{uri}">
        <z:itemType>document</z:itemType>
        <dc:title>{title}</dc:title>
        <dc:creator>
            <foaf:Person>
                <foaf:surname>PDF Analysis</foaf:surname>
                <foaf:givenName>Extracted</foaf:givenName>
            </foaf:Person>
        </dc:creator>
        <dc:date>{date}</dc:date>
        <dc:description>{abstract}</dc:description>
        <z:libraryCatalog>PDF Knowledge Extractor</z:libraryCatalog>
        <dc:subject>{keywords}</dc:subject>
    </bib:Document>""",
            'collection': """
    <z:Collection rdf:about="{collection_uri}">
        <dc:title>{collection_title}</dc:title>
        <z:hasMember rdf:resource="{member_uri}"/>
    </z:Collection>""",
            'note': """
    <bib:Memo rdf:about="{note_uri}">
        <rdf:value>{note_content}</rdf:value>
        <z:itemType>note</z:itemType>
    </bib:Memo>"""
        }
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to Zotero RDF format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create RDF XML structure
            rdf_content = self.render_template('rdf_header', {})
            
            exported_docs = []
            total_concepts = 0
            
            # Create Zotero items for each document
            for doc_id, doc_text in documents.items():
                doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
                semantic_data = analysis_data.get('semantic_analysis', {})
                
                # Get concepts for keywords
                doc_concepts = []
                if semantic_data and semantic_data.get('concepts'):
                    doc_concepts = [c for c in semantic_data['concepts'] 
                                  if doc_id in c.get('document_ids', [])]
                    doc_concepts = self.filter_concepts(doc_concepts, doc_id)
                
                # Create keywords from concepts
                keywords = ', '.join([c.get('text', '') for c in doc_concepts[:10]])
                
                # Generate abstract/description
                abstract = self._generate_abstract(doc_text, doc_analysis)
                
                # Create document item
                doc_uri = f"#document_{hash(doc_id)}"
                doc_item = self.render_template('document_item', {
                    'uri': doc_uri,
                    'title': self._escape_xml(doc_id),
                    'date': datetime.now().strftime('%Y-%m-%d'),
                    'abstract': self._escape_xml(abstract),
                    'keywords': self._escape_xml(keywords)
                })
                
                rdf_content += doc_item
                
                # Create notes with concepts and relationships
                if self.config.include_concepts and doc_concepts:
                    note_content = self._create_concept_note(doc_concepts)
                    note_uri = f"#note_concepts_{hash(doc_id)}"
                    note_item = self.render_template('note', {
                        'note_uri': note_uri,
                        'note_content': self._escape_xml(note_content)
                    })
                    rdf_content += note_item
                
                exported_docs.append(doc_id)
                total_concepts += len(doc_concepts)
            
            # Create collections for clusters
            if self.config.include_clusters and analysis_data.get('semantic_analysis', {}).get('clusters'):
                for cluster in analysis_data['semantic_analysis']['clusters']:
                    collection_uri = f"#collection_{cluster.get('cluster_id')}"
                    collection_item = self.render_template('collection', {
                        'collection_uri': collection_uri,
                        'collection_title': self._escape_xml(cluster.get('cluster_label', 'Unnamed Cluster')),
                        'member_uri': f"#document_{hash(cluster.get('document_ids', [])[0])}"
                    })
                    rdf_content += collection_item
            
            rdf_content += "\n</rdf:RDF>"
            
            # Write RDF file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(rdf_content)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                stats={'zotero_items': len(exported_docs)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Zotero export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _generate_abstract(self, doc_text: str, analysis: Dict) -> str:
        """Generate abstract from document."""
        sentences = doc_text.split('.')[:3]
        abstract = '. '.join(sentences).strip()
        return abstract[:500] + "..." if len(abstract) > 500 else abstract
    
    def _create_concept_note(self, concepts: List[Dict]) -> str:
        """Create a note with concept information."""
        note = "Key Concepts:\n\n"
        for concept in concepts:
            note += f"• {concept.get('text', '')} (importance: {concept.get('importance_score', 0):.2f})\n"
            if concept.get('context_sentences'):
                note += f"  Context: {concept['context_sentences'][0][:100]}...\n"
            note += "\n"
        return note
    
    def _escape_xml(self, text: str) -> str:
        """Escape XML special characters."""
        if not text:
            return ""
        return (text.replace('&', '&amp;')
                   .replace('<', '&lt;')
                   .replace('>', '&gt;')
                   .replace('"', '&quot;')
                   .replace("'", '&apos;'))


class LaTeXExporter(TemplateExporter):
    """Export to LaTeX document with citations and bibliography."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['latex', 'tex']
    
    @property
    def file_extension(self) -> str:
        return '.tex'
    
    @property
    def default_templates(self) -> Dict[str, str]:
        return {
            'document_template': """\\documentclass[12pt,a4paper]{{article}}
\\usepackage[utf8]{{inputenc}}
\\usepackage[T1]{{fontenc}}
\\usepackage{{amsmath,amsfonts,amssymb}}
\\usepackage{{graphicx}}
\\usepackage{{hyperref}}
\\usepackage{{natbib}}
\\usepackage{{booktabs}}
\\usepackage{{longtable}}

\\title{{PDF Knowledge Analysis Report}}
\\author{{PDF Knowledge Extractor v2.2}}
\\date{{{date}}}

\\begin{{document}}

\\maketitle

\\begin{{abstract}}
This report presents an analysis of {total_documents} PDF documents using semantic analysis and knowledge extraction techniques. The analysis identified {total_concepts} concepts and {total_relationships} relationships across the document collection.
\\end{{abstract}}

\\tableofcontents

\\section{{Executive Summary}}
{executive_summary}

\\section{{Document Analysis}}
{document_sections}

\\section{{Semantic Analysis Results}}
{semantic_analysis_section}

\\section{{Concept Index}}
{concept_index}

\\section{{Document Relationships}}
{relationships_section}

\\bibliography{{references}}
\\bibliographystyle{{plain}}

\\end{{document}}""",
            'document_section': """\\subsection{{{title}}}
\\label{{doc:{label}}}

\\begin{{itemize}}
    \\item \\textbf{{Word Count:}} {word_count}
    \\item \\textbf{{Concept Count:}} {concept_count}
    \\item \\textbf{{Analysis Date:}} {date}
\\end{{itemize}}

\\subsubsection{{Summary}}
{summary}

\\subsubsection{{Key Concepts}}
\\begin{{itemize}}
{concepts_list}
\\end{{itemize}}

\\subsubsection{{Related Documents}}
{related_docs}
""",
            'concept_item': "    \\item \\texttt{{{concept_text}}} (importance: {importance:.2f})",
            'bibliography_entry': "@misc{{{cite_key},\n  title={{{title}}},\n  author={{PDF Knowledge Extractor}},\n  year={{{year}}},\n  note={{Analyzed document}}\n}}"
        }
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to LaTeX document format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Generate LaTeX content sections
            document_sections = ""
            concept_index = ""
            relationships_section = ""
            executive_summary = self._generate_executive_summary(analysis_data, documents)
            
            exported_docs = []
            total_concepts = 0
            total_relationships = 0
            
            # Create bibliography file
            bib_path = output_path.with_suffix('.bib')
            bib_entries = []
            
            # Process each document
            for doc_id, doc_text in documents.items():
                doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
                semantic_data = analysis_data.get('semantic_analysis', {})
                
                # Get concepts
                doc_concepts = []
                if semantic_data and semantic_data.get('concepts'):
                    doc_concepts = [c for c in semantic_data['concepts'] 
                                  if doc_id in c.get('document_ids', [])]
                    doc_concepts = self.filter_concepts(doc_concepts, doc_id)
                
                # Create document section
                doc_section = self._create_latex_document_section(doc_id, doc_text, doc_concepts, doc_analysis, semantic_data)
                document_sections += doc_section + "\n\n"
                
                # Add to bibliography
                cite_key = self._create_cite_key(doc_id)
                bib_entry = self.render_template('bibliography_entry', {
                    'cite_key': cite_key,
                    'title': self._escape_latex(doc_id),
                    'year': datetime.now().year
                })
                bib_entries.append(bib_entry)
                
                exported_docs.append(doc_id)
                total_concepts += len(doc_concepts)
            
            # Create concept index
            if semantic_data and semantic_data.get('concepts'):
                concept_index = self._create_concept_index_latex(semantic_data['concepts'])
            
            # Create relationships section
            if semantic_data and semantic_data.get('similarities'):
                relationships_section = self._create_relationships_section_latex(semantic_data['similarities'])
                total_relationships = len(semantic_data['similarities'])
            
            # Create semantic analysis section
            semantic_section = self._create_semantic_analysis_section(analysis_data.get('semantic_analysis', {}))
            
            # Generate main LaTeX document
            latex_content = self.render_template('document_template', {
                'date': datetime.now().strftime('%Y-%m-%d'),
                'total_documents': len(documents),
                'total_concepts': total_concepts,
                'total_relationships': total_relationships,
                'executive_summary': executive_summary,
                'document_sections': document_sections,
                'semantic_analysis_section': semantic_section,
                'concept_index': concept_index,
                'relationships_section': relationships_section
            })
            
            # Write LaTeX file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(latex_content)
            
            # Write bibliography file
            with open(bib_path, 'w', encoding='utf-8') as f:
                f.write('\n\n'.join(bib_entries))
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                exported_relationships=total_relationships,
                stats={'latex_sections': len(exported_docs), 'bibliography_entries': len(bib_entries)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"LaTeX export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_latex_document_section(self, doc_id: str, doc_text: str, concepts: List[Dict], 
                                     doc_analysis: Dict, semantic_data: Dict) -> str:
        """Create LaTeX section for a document."""
        
        # Build concepts list
        concepts_list = ""
        for concept in concepts:
            concept_item = self.render_template('concept_item', {
                'concept_text': self._escape_latex(concept.get('text', '')),
                'importance': concept.get('importance_score', 0)
            })
            concepts_list += concept_item + "\n"
        
        # Build related documents
        related_docs = "\\begin{itemize}\n"
        if semantic_data and semantic_data.get('similarities'):
            for sim in semantic_data['similarities']:
                if (sim.get('doc1_id') == doc_id or sim.get('doc2_id') == doc_id) and \
                   sim.get('similarity_score', 0) >= self.config.min_similarity_score:
                    other_doc = sim.get('doc2_id') if sim.get('doc1_id') == doc_id else sim.get('doc1_id')
                    cite_key = self._create_cite_key(other_doc)
                    related_docs += f"    \\item \\cite{{{cite_key}}} (similarity: {sim.get('similarity_score', 0):.2f})\n"
        related_docs += "\\end{itemize}"
        
        # Generate summary
        summary = self._generate_summary(doc_text, doc_analysis)
        
        return self.render_template('document_section', {
            'title': self._escape_latex(doc_id),
            'label': self._create_label(doc_id),
            'word_count': doc_analysis.get('word_count', 0),
            'concept_count': len(concepts),
            'date': datetime.now().strftime('%Y-%m-%d'),
            'summary': self._escape_latex(summary),
            'concepts_list': concepts_list,
            'related_docs': related_docs
        })
    
    def _create_concept_index_latex(self, concepts: List[Dict]) -> str:
        """Create LaTeX concept index."""
        content = "\\begin{longtable}{lrp{8cm}}\n"
        content += "\\toprule\n"
        content += "Concept & Importance & Context \\\\\n"
        content += "\\midrule\n"
        
        # Sort concepts by importance
        sorted_concepts = sorted(concepts, key=lambda x: x.get('importance_score', 0), reverse=True)
        
        for concept in sorted_concepts[:50]:  # Top 50 concepts
            context = concept.get('context_sentences', [''])[0][:100] if concept.get('context_sentences') else ''
            content += f"\\texttt{{{self._escape_latex(concept.get('text', ''))}}} & " \
                      f"{concept.get('importance_score', 0):.2f} & " \
                      f"{self._escape_latex(context)}... \\\\\n"
        
        content += "\\bottomrule\n"
        content += "\\end{longtable}\n"
        return content
    
    def _create_relationships_section_latex(self, similarities: List[Dict]) -> str:
        """Create LaTeX relationships section."""
        content = "\\begin{longtable}{llr}\n"
        content += "\\toprule\n"
        content += "Document 1 & Document 2 & Similarity \\\\\n"
        content += "\\midrule\n"
        
        # Sort by similarity score
        sorted_sims = sorted(similarities, key=lambda x: x.get('similarity_score', 0), reverse=True)
        
        for sim in sorted_sims[:30]:  # Top 30 relationships
            doc1_key = self._create_cite_key(sim.get('doc1_id', ''))
            doc2_key = self._create_cite_key(sim.get('doc2_id', ''))
            content += f"\\cite{{{doc1_key}}} & \\cite{{{doc2_key}}} & {sim.get('similarity_score', 0):.3f} \\\\\n"
        
        content += "\\bottomrule\n"
        content += "\\end{longtable}\n"
        return content
    
    def _create_semantic_analysis_section(self, semantic_data: Dict) -> str:
        """Create semantic analysis section."""
        content = "This section presents the results of semantic analysis performed on the document collection.\n\n"
        
        if semantic_data.get('clusters'):
            content += "\\subsection{Document Clusters}\n"
            content += f"The analysis identified {len(semantic_data['clusters'])} distinct document clusters:\n\n"
            content += "\\begin{itemize}\n"
            
            for cluster in semantic_data['clusters']:
                content += f"\\item \\textbf{{{self._escape_latex(cluster.get('cluster_label', 'Unnamed'))}}}: "
                content += f"{len(cluster.get('document_ids', []))} documents "
                content += f"(coherence: {cluster.get('coherence_score', 0):.2f})\n"
            
            content += "\\end{itemize}\n\n"
        
        return content
    
    def _generate_executive_summary(self, analysis_data: Dict, documents: Dict) -> str:
        """Generate executive summary."""
        summary = f"This report analyzes {len(documents)} PDF documents using advanced semantic analysis techniques. "
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        if semantic_data:
            if semantic_data.get('concepts'):
                summary += f"A total of {len(semantic_data['concepts'])} concepts were extracted across all documents. "
            
            if semantic_data.get('clusters'):
                summary += f"Documents were organized into {len(semantic_data['clusters'])} thematic clusters. "
            
            if semantic_data.get('similarities'):
                summary += f"The analysis identified {len(semantic_data['similarities'])} significant relationships between documents."
        
        return summary
    
    def _generate_summary(self, doc_text: str, analysis: Dict) -> str:
        """Generate document summary."""
        sentences = doc_text.split('.')[:2]
        summary = '. '.join(sentences).strip()
        return summary[:400] + "..." if len(summary) > 400 else summary
    
    def _escape_latex(self, text: str) -> str:
        """Escape LaTeX special characters."""
        if not text:
            return ""
        
        # LaTeX special characters
        replacements = {
            '\\': '\\textbackslash{}',
            '{': '\\{',
            '}': '\\}',
            '$': '\\$',
            '&': '\\&',
            '%': '\\%',
            '#': '\\#',
            '^': '\\textasciicircum{}',
            '_': '\\_',
            '~': '\\textasciitilde{}'
        }
        
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        return text
    
    def _create_cite_key(self, doc_id: str) -> str:
        """Create a valid BibTeX citation key."""
        key = re.sub(r'[^\w\-]', '', doc_id.replace(' ', '_'))
        return key[:50]  # Limit length
    
    def _create_label(self, doc_id: str) -> str:
        """Create a valid LaTeX label."""
        label = re.sub(r'[^\w\-]', '', doc_id.replace(' ', '_'))
        return label[:50]


class GoogleDocsExporter(TemplateExporter):
    """Export to Google Docs compatible HTML format."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['gdocs', 'html']
    
    @property
    def file_extension(self) -> str:
        return '.html'
    
    @property
    def default_templates(self) -> Dict[str, str]:
        return {
            'html_template': """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>PDF Knowledge Analysis Report</title>
    <style>
        body {{ font-family: 'Times New Roman', serif; margin: 1in; line-height: 1.6; }}
        h1 {{ color: #1a73e8; font-size: 24pt; }}
        h2 {{ color: #1a73e8; font-size: 18pt; border-bottom: 2px solid #e8f0fe; }}
        h3 {{ color: #1a73e8; font-size: 14pt; }}
        .concept {{ background-color: #e8f0fe; padding: 4px 8px; border-radius: 4px; }}
        .importance {{ font-weight: bold; color: #137333; }}
        .similarity {{ font-weight: bold; color: #ea4335; }}
        .metadata {{ background-color: #f8f9fa; padding: 10px; border-left: 4px solid #1a73e8; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #e8f0fe; }}
    </style>
</head>
<body>
    <h1>PDF Knowledge Analysis Report</h1>
    
    <div class="metadata">
        <p><strong>Generated:</strong> {timestamp}</p>
        <p><strong>Total Documents:</strong> {total_documents}</p>
        <p><strong>Total Concepts:</strong> {total_concepts}</p>
        <p><strong>Total Relationships:</strong> {total_relationships}</p>
    </div>

    <h2>Executive Summary</h2>
    <p>{executive_summary}</p>

    <h2>Document Analysis</h2>
    {document_sections}

    <h2>Concept Overview</h2>
    {concept_overview}

    <h2>Document Relationships</h2>
    {relationships_section}

    <footer>
        <p><em>Generated by PDF Knowledge Extractor v2.2</em></p>
    </footer>
</body>
</html>""",
            'document_section': """<h3>{title}</h3>
<div class="metadata">
    <p><strong>Word Count:</strong> {word_count} | <strong>Concepts:</strong> {concept_count}</p>
</div>
<p><strong>Summary:</strong> {summary}</p>
<p><strong>Key Concepts:</strong></p>
<ul>
{concepts_list}
</ul>
<p><strong>Related Documents:</strong></p>
<ul>
{related_docs}
</ul>""",
            'concept_item': '<li><span class="concept">{concept_text}</span> <span class="importance">({importance:.2f})</span></li>',
            'related_item': '<li>{doc_title} <span class="similarity">(similarity: {similarity:.2f})</span></li>'
        }
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to Google Docs compatible HTML format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            document_sections = ""
            exported_docs = []
            total_concepts = 0
            total_relationships = 0
            
            # Process each document
            for doc_id, doc_text in documents.items():
                doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
                semantic_data = analysis_data.get('semantic_analysis', {})
                
                # Get concepts
                doc_concepts = []
                if semantic_data and semantic_data.get('concepts'):
                    doc_concepts = [c for c in semantic_data['concepts'] 
                                  if doc_id in c.get('document_ids', [])]
                    doc_concepts = self.filter_concepts(doc_concepts, doc_id)
                
                # Create document section
                doc_section = self._create_html_document_section(doc_id, doc_text, doc_concepts, doc_analysis, semantic_data)
                document_sections += doc_section + "\n"
                
                exported_docs.append(doc_id)
                total_concepts += len(doc_concepts)
            
            # Create concept overview
            concept_overview = self._create_concept_overview_html(analysis_data.get('semantic_analysis', {}).get('concepts', []))
            
            # Create relationships section
            relationships_section = self._create_relationships_html(analysis_data.get('semantic_analysis', {}).get('similarities', []))
            if analysis_data.get('semantic_analysis', {}).get('similarities'):
                total_relationships = len(analysis_data['semantic_analysis']['similarities'])
            
            # Generate executive summary
            executive_summary = self._generate_executive_summary(analysis_data, documents)
            
            # Create HTML document
            html_content = self.render_template('html_template', {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_documents': len(documents),
                'total_concepts': total_concepts,
                'total_relationships': total_relationships,
                'executive_summary': executive_summary,
                'document_sections': document_sections,
                'concept_overview': concept_overview,
                'relationships_section': relationships_section
            })
            
            # Write HTML file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                exported_relationships=total_relationships,
                stats={'html_sections': len(exported_docs)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Google Docs HTML export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_html_document_section(self, doc_id: str, doc_text: str, concepts: List[Dict], 
                                    doc_analysis: Dict, semantic_data: Dict) -> str:
        """Create HTML section for a document."""
        
        # Build concepts list
        concepts_list = ""
        for concept in concepts:
            concept_item = self.render_template('concept_item', {
                'concept_text': self._escape_html(concept.get('text', '')),
                'importance': concept.get('importance_score', 0)
            })
            concepts_list += concept_item + "\n"
        
        # Build related documents
        related_docs = ""
        if semantic_data and semantic_data.get('similarities'):
            for sim in semantic_data['similarities']:
                if (sim.get('doc1_id') == doc_id or sim.get('doc2_id') == doc_id) and \
                   sim.get('similarity_score', 0) >= self.config.min_similarity_score:
                    other_doc = sim.get('doc2_id') if sim.get('doc1_id') == doc_id else sim.get('doc1_id')
                    related_item = self.render_template('related_item', {
                        'doc_title': self._escape_html(other_doc),
                        'similarity': sim.get('similarity_score', 0)
                    })
                    related_docs += related_item + "\n"
        
        if not related_docs:
            related_docs = "<li>No related documents found.</li>"
        
        # Generate summary
        summary = self._generate_summary(doc_text, doc_analysis)
        
        return self.render_template('document_section', {
            'title': self._escape_html(doc_id),
            'word_count': doc_analysis.get('word_count', 0),
            'concept_count': len(concepts),
            'summary': self._escape_html(summary),
            'concepts_list': concepts_list,
            'related_docs': related_docs
        })
    
    def _create_concept_overview_html(self, concepts: List[Dict]) -> str:
        """Create HTML concept overview table."""
        if not concepts:
            return "<p>No concepts extracted.</p>"
        
        html = "<table>\n<tr><th>Concept</th><th>Type</th><th>Importance</th><th>Documents</th></tr>\n"
        
        # Sort by importance
        sorted_concepts = sorted(concepts, key=lambda x: x.get('importance_score', 0), reverse=True)
        
        for concept in sorted_concepts[:30]:  # Top 30 concepts
            html += "<tr>\n"
            html += f"<td>{self._escape_html(concept.get('text', ''))}</td>\n"
            html += f"<td>{concept.get('concept_type', 'unknown')}</td>\n"
            html += f"<td>{concept.get('importance_score', 0):.3f}</td>\n"
            html += f"<td>{len(concept.get('document_ids', []))}</td>\n"
            html += "</tr>\n"
        
        html += "</table>"
        return html
    
    def _create_relationships_html(self, similarities: List[Dict]) -> str:
        """Create HTML relationships section."""
        if not similarities:
            return "<p>No document relationships identified.</p>"
        
        html = "<table>\n<tr><th>Document 1</th><th>Document 2</th><th>Similarity</th></tr>\n"
        
        # Sort by similarity
        sorted_sims = sorted(similarities, key=lambda x: x.get('similarity_score', 0), reverse=True)
        
        for sim in sorted_sims[:20]:  # Top 20 relationships
            html += "<tr>\n"
            html += f"<td>{self._escape_html(sim.get('doc1_id', ''))}</td>\n"
            html += f"<td>{self._escape_html(sim.get('doc2_id', ''))}</td>\n"
            html += f"<td>{sim.get('similarity_score', 0):.3f}</td>\n"
            html += "</tr>\n"
        
        html += "</table>"
        return html
    
    def _generate_executive_summary(self, analysis_data: Dict, documents: Dict) -> str:
        """Generate executive summary."""
        summary = f"This report analyzes {len(documents)} PDF documents using semantic analysis. "
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        if semantic_data:
            if semantic_data.get('concepts'):
                summary += f"The analysis extracted {len(semantic_data['concepts'])} unique concepts "
            
            if semantic_data.get('clusters'):
                summary += f"and organized documents into {len(semantic_data['clusters'])} thematic clusters. "
            
            if semantic_data.get('similarities'):
                summary += f"A total of {len(semantic_data['similarities'])} relationships were identified between documents."
        
        return summary
    
    def _generate_summary(self, doc_text: str, analysis: Dict) -> str:
        """Generate document summary."""
        sentences = doc_text.split('.')[:2]
        summary = '. '.join(sentences).strip()
        return summary[:300] + "..." if len(summary) > 300 else summary
    
    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        if not text:
            return ""
        return (text.replace('&', '&amp;')
                   .replace('<', '&lt;')
                   .replace('>', '&gt;')
                   .replace('"', '&quot;')
                   .replace("'", '&#39;'))


class CSVExporter(BaseExporter):
    """Export to CSV format for data analysis."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['csv']
    
    @property
    def file_extension(self) -> str:
        return '.csv'
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to CSV format."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create comprehensive CSV with all data
            csv_data = []
            exported_docs = []
            total_concepts = 0
            
            # CSV headers
            headers = ['document_id', 'word_count', 'character_count', 'concept_text', 
                      'concept_type', 'concept_importance', 'concept_frequency',
                      'related_document', 'similarity_score']
            
            semantic_data = analysis_data.get('semantic_analysis', {})
            
            # Process each document
            for doc_id, doc_text in documents.items():
                doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
                
                # Get concepts for this document
                doc_concepts = []
                if semantic_data and semantic_data.get('concepts'):
                    doc_concepts = [c for c in semantic_data['concepts'] 
                                  if doc_id in c.get('document_ids', [])]
                    doc_concepts = self.filter_concepts(doc_concepts, doc_id)
                
                # Get related documents
                related_docs = []
                if semantic_data and semantic_data.get('similarities'):
                    for sim in semantic_data['similarities']:
                        if (sim.get('doc1_id') == doc_id or sim.get('doc2_id') == doc_id) and \
                           sim.get('similarity_score', 0) >= self.config.min_similarity_score:
                            other_doc = sim.get('doc2_id') if sim.get('doc1_id') == doc_id else sim.get('doc1_id')
                            related_docs.append({
                                'doc_id': other_doc,
                                'similarity': sim.get('similarity_score', 0)
                            })
                
                # Create rows for each concept
                if doc_concepts:
                    for concept in doc_concepts:
                        for related in related_docs or [{}]:
                            csv_data.append([
                                doc_id,
                                doc_analysis.get('word_count', 0),
                                doc_analysis.get('character_count', 0),
                                concept.get('text', ''),
                                concept.get('concept_type', ''),
                                concept.get('importance_score', 0),
                                concept.get('frequency', 0),
                                related.get('doc_id', ''),
                                related.get('similarity', 0)
                            ])
                else:
                    # Document without concepts
                    for related in related_docs or [{}]:
                        csv_data.append([
                            doc_id,
                            doc_analysis.get('word_count', 0),
                            doc_analysis.get('character_count', 0),
                            '', '', 0, 0,
                            related.get('doc_id', ''),
                            related.get('similarity', 0)
                        ])
                
                exported_docs.append(doc_id)
                total_concepts += len(doc_concepts)
            
            # Write CSV file
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                writer.writerows(csv_data)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                stats={'csv_rows': len(csv_data)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"CSV export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )


class ExcelExporter(BaseExporter):
    """Export to Excel format with multiple sheets and formatting."""
    
    @property
    def supported_formats(self) -> List[str]:
        return ['excel', 'xlsx']
    
    @property
    def file_extension(self) -> str:
        return '.xlsx'
    
    def export(self, analysis_data: Dict[str, Any], documents: Dict[str, str]) -> ExportResult:
        """Export to Excel format with multiple sheets."""
        start_time = time.time()
        
        try:
            errors = self.validate_config()
            if errors:
                return self.create_export_result(False, self.config.output_path, [], 0, errors=errors)
            
            output_path = self.prepare_output_path()
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            exported_docs = []
            total_concepts = 0
            total_relationships = 0
            
            semantic_data = analysis_data.get('semantic_analysis', {})
            
            # Create Documents sheet
            doc_sheet = workbook.create_sheet("Documents")
            self._create_documents_sheet(doc_sheet, analysis_data, documents)
            
            # Create Concepts sheet
            if semantic_data.get('concepts'):
                concepts_sheet = workbook.create_sheet("Concepts")
                self._create_concepts_sheet(concepts_sheet, semantic_data['concepts'])
                total_concepts = len(semantic_data['concepts'])
            
            # Create Relationships sheet
            if semantic_data.get('similarities'):
                rel_sheet = workbook.create_sheet("Relationships")
                self._create_relationships_sheet(rel_sheet, semantic_data['similarities'])
                total_relationships = len(semantic_data['similarities'])
            
            # Create Clusters sheet
            if semantic_data.get('clusters'):
                cluster_sheet = workbook.create_sheet("Clusters")
                self._create_clusters_sheet(cluster_sheet, semantic_data['clusters'])
            
            # Create Summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, analysis_data, documents)
            
            # Make Summary the first sheet
            workbook.move_sheet(summary_sheet, offset=-len(workbook.worksheets))
            
            exported_docs = list(documents.keys())
            
            # Save Excel file
            workbook.save(output_path)
            
            execution_time = time.time() - start_time
            
            return self.create_export_result(
                success=True,
                output_path=output_path,
                exported_docs=exported_docs,
                execution_time=execution_time,
                exported_concepts=total_concepts,
                exported_relationships=total_relationships,
                stats={'excel_sheets': len(workbook.worksheets)}
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Excel export failed: {e}")
            return self.create_export_result(
                False, self.config.output_path, [], execution_time, errors=[str(e)]
            )
    
    def _create_documents_sheet(self, sheet, analysis_data: Dict, documents: Dict):
        """Create documents overview sheet."""
        # Headers
        headers = ['Document ID', 'Word Count', 'Character Count', 'Concepts Count', 'Topics', 'Sentiment']
        sheet.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for doc_id, doc_text in documents.items():
            doc_analysis = analysis_data.get('individual_analyses', {}).get(doc_id, {})
            
            # Count concepts for this document
            concept_count = 0
            semantic_data = analysis_data.get('semantic_analysis', {})
            if semantic_data and semantic_data.get('concepts'):
                concept_count = len([c for c in semantic_data['concepts'] 
                                   if doc_id in c.get('document_ids', [])])
            
            # Get top topics
            topics = doc_analysis.get('topics', [])
            top_topics = ', '.join([t.get('topic', '') for t in topics[:3]])
            
            # Get sentiment
            sentiment = doc_analysis.get('sentiment', {})
            sentiment_str = f"{sentiment.get('sentiment', 'neutral')} ({sentiment.get('score', 0):.2f})"
            
            sheet.append([
                doc_id,
                doc_analysis.get('word_count', 0),
                doc_analysis.get('character_count', 0),
                concept_count,
                top_topics,
                sentiment_str
            ])
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_concepts_sheet(self, sheet, concepts: List[Dict]):
        """Create concepts sheet."""
        headers = ['Concept', 'Type', 'Importance', 'Frequency', 'Document Count', 'Documents']
        sheet.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Sort concepts by importance
        sorted_concepts = sorted(concepts, key=lambda x: x.get('importance_score', 0), reverse=True)
        
        # Data rows
        for concept in sorted_concepts:
            doc_ids = concept.get('document_ids', [])
            doc_list = ', '.join(doc_ids[:3])  # Show first 3 documents
            if len(doc_ids) > 3:
                doc_list += f" (+{len(doc_ids) - 3} more)"
            
            sheet.append([
                concept.get('text', ''),
                concept.get('concept_type', ''),
                concept.get('importance_score', 0),
                concept.get('frequency', 0),
                len(doc_ids),
                doc_list
            ])
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_relationships_sheet(self, sheet, similarities: List[Dict]):
        """Create relationships sheet."""
        headers = ['Document 1', 'Document 2', 'Similarity Score', 'Shared Concepts']
        sheet.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Sort by similarity score
        sorted_sims = sorted(similarities, key=lambda x: x.get('similarity_score', 0), reverse=True)
        
        # Data rows
        for sim in sorted_sims:
            shared_concepts = ', '.join(sim.get('shared_concepts', [])[:5])  # First 5
            
            sheet.append([
                sim.get('doc1_id', ''),
                sim.get('doc2_id', ''),
                sim.get('similarity_score', 0),
                shared_concepts
            ])
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_clusters_sheet(self, sheet, clusters: List[Dict]):
        """Create clusters sheet."""
        headers = ['Cluster ID', 'Label', 'Document Count', 'Coherence Score', 'Main Topics', 'Documents']
        sheet.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for cluster in clusters:
            doc_ids = cluster.get('document_ids', [])
            doc_list = ', '.join(doc_ids[:3])  # Show first 3
            if len(doc_ids) > 3:
                doc_list += f" (+{len(doc_ids) - 3} more)"
            
            topics = ', '.join(cluster.get('main_topics', [])[:5])
            
            sheet.append([
                cluster.get('cluster_id', ''),
                cluster.get('cluster_label', ''),
                len(doc_ids),
                cluster.get('coherence_score', 0),
                topics,
                doc_list
            ])
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, sheet, analysis_data: Dict, documents: Dict):
        """Create summary overview sheet."""
        # Title
        sheet['A1'] = 'PDF Knowledge Analysis Summary'
        sheet['A1'].font = Font(size=16, bold=True, color="2E86AB")
        
        # Basic statistics
        row = 3
        stats = [
            ('Total Documents', len(documents)),
            ('Total Words', sum(analysis_data.get('individual_analyses', {}).get(doc_id, {}).get('word_count', 0) 
                              for doc_id in documents.keys())),
            ('Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        semantic_data = analysis_data.get('semantic_analysis', {})
        if semantic_data:
            if semantic_data.get('concepts'):
                stats.append(('Total Concepts', len(semantic_data['concepts'])))
            if semantic_data.get('similarities'):
                stats.append(('Document Relationships', len(semantic_data['similarities'])))
            if semantic_data.get('clusters'):
                stats.append(('Document Clusters', len(semantic_data['clusters'])))
        
        for stat_name, stat_value in stats:
            sheet[f'A{row}'] = stat_name
            sheet[f'B{row}'] = stat_value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15